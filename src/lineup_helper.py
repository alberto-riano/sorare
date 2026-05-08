#!/usr/bin/env python3
# ============================================================
# CONFIGURACIÓN
# ============================================================
EXCEL_PATH = "../input/lineups.xlsx"
ACTUALIZAR_CARTAS = False     # True = descargar cartas de nuevo desde Sorare
MOSTRAR_DETALLE = False       # True = mostrar listado de jugadores por posición
MAX_PUNTOS = 260              # Máximo de puntos por alineación
PESO_CUOTAS = 0.3             # Peso de las cuotas (0 = solo media, 0.5 = mucha influencia)
PESO_GOLES = 0.2              # Peso de over/under y clean sheet (0 = ignorar)
BONUS_POR_DEF = 3             # Bonus por combinar POR+DEF del mismo equipo
# ============================================================
"""
Lee el Excel de alineaciones (última pestaña) y muestra los jugadores
organizados por posición (porteros, defensas, medios, delanteros),
distinguiendo entre in-season y classic.

Estructura esperada del Excel:
  - 4 bloques de 2 columnas (nombre + media): POR, DEF, MED, DEL
  - Una fila completamente vacía en un bloque separa in-season de classic
"""
import os
import sys
import json
import openpyxl
from difflib import SequenceMatcher
from itertools import combinations
from collections import Counter
import requests
from datetime import datetime, timezone

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sorare_utils import graphql_request, build_headers

CARDS_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output', 'mis_cartas.json')
ODDS_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output', 'odds_cache.json')

POSITIONS = ["Porteros", "Defensas", "Medios", "Delanteros"]
# Cada posición ocupa 2 columnas (nombre + media), empezando en col_start
# Se detectan automáticamente al leer la hoja


def find_column_pairs(ws):
    """
    Detecta los pares de columnas (nombre, media) buscando las primeras
    filas con datos. Devuelve lista de tuplas (col_nombre, col_media).
    """
    # Buscar todas las columnas que tienen al menos un dato
    cols_with_data = set()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cols_with_data.add(cell.column)

    if not cols_with_data:
        return []

    # Ordenar columnas y agrupar en pares consecutivos
    sorted_cols = sorted(cols_with_data)
    pairs = []
    i = 0
    while i < len(sorted_cols) - 1:
        col_name = sorted_cols[i]
        col_score = sorted_cols[i + 1]
        # Verificar que la segunda columna tiene números (medias)
        has_numbers = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=col_score, max_col=col_score):
            if row[0].value is not None and isinstance(row[0].value, (int, float)):
                has_numbers = True
                break
        if has_numbers:
            pairs.append((col_name, col_score))
            i += 2
        else:
            i += 1

    return pairs


def read_position_column(ws, col_name, col_score):
    """
    Lee una columna de posición y devuelve (in_season, classic).
    Cada uno es una lista de (nombre, media).
    La separación ocurre cuando hay una fila vacía DESPUÉS de haber
    encontrado al menos un jugador.
    """
    in_season = []
    classic = []
    current = in_season
    found_any = False
    blank_streak = 0

    for row_num in range(1, ws.max_row + 1):
        name = ws.cell(row=row_num, column=col_name).value
        score = ws.cell(row=row_num, column=col_score).value

        if name is not None and score is not None:
            if found_any and blank_streak > 0:
                # Hubo un hueco → lo que viene ahora es classic
                current = classic
            current.append((str(name).strip(), int(score)))
            found_any = True
            blank_streak = 0
        else:
            if found_any:
                blank_streak += 1

    return in_season, classic


def print_position(pos_name, in_season, classic, team_map):
    """Imprime los jugadores de una posición."""
    print(f"\n{'─' * 55}")
    print(f"  {pos_name.upper()}")
    print(f"{'─' * 55}")

    if in_season:
        print(f"  In-Season ({len(in_season)}):")
        for name, score in in_season:
            team = team_map.get(name, '?')
            print(f"    {name:<20} {score:>3}   {team}")
    else:
        print("  In-Season: (ninguno)")

    if classic:
        print(f"  Classic ({len(classic)}):")
        for name, score in classic:
            team = team_map.get(name, '?')
            print(f"    {name:<20} {score:>3}   {team}")


def load_lineup(excel_path=None):
    """
    Carga la última pestaña del Excel y devuelve un dict:
    { posición: { 'in_season': [(nombre, media), ...], 'classic': [...] } }
    """
    path = excel_path or EXCEL_PATH
    if not os.path.isabs(path):
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)

    if not os.path.exists(path):
        print(f"❌ No se encuentra el fichero: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[-1]]

    print(f"📄 Fichero: {os.path.basename(path)}")
    print(f"📑 Pestaña: {ws.title}")

    pairs = find_column_pairs(ws)
    if len(pairs) < 4:
        print(f"⚠️  Se esperaban 4 pares de columnas, se encontraron {len(pairs)}")
        print(f"   Pares detectados: {pairs}")

    lineup = {}
    for i, (col_name, col_score) in enumerate(pairs[:4]):
        pos = POSITIONS[i] if i < len(POSITIONS) else f"Posición {i + 1}"
        in_season, classic = read_position_column(ws, col_name, col_score)
        lineup[pos] = {'in_season': in_season, 'classic': classic}

    return lineup


def fetch_my_rare_cards(headers):
    """Descarga todas las cartas rare del usuario con nombre y equipo."""
    query = '''
    query GetMyCards($first: Int!, $after: String) {
      currentUser {
        cards(first: $first, after: $after) {
          pageInfo { hasNextPage endCursor }
          nodes {
            rarityTyped
            anyPlayer { displayName }
            anyTeam { name }
          }
        }
      }
    }
    '''
    cards = []
    cursor = None
    page = 0
    while True:
        page += 1
        data = graphql_request(query, {'first': 50, 'after': cursor}, headers=headers)
        page_data = data['currentUser']['cards']
        for node in page_data['nodes']:
            if node['rarityTyped'] in ('rare', 'classic'):
                player = node.get('anyPlayer') or {}
                team = node.get('anyTeam') or {}
                cards.append({
                    'name': player.get('displayName', ''),
                    'team': team.get('name', ''),
                    'rarity': node['rarityTyped'],
                })
        if not page_data['pageInfo']['hasNextPage']:
            break
        cursor = page_data['pageInfo']['endCursor']
        print(f"\r  Cargando cartas... página {page}", end='', flush=True)
    print(f"\r  Cargadas {len(cards)} cartas rare/classic" + " " * 20)

    # Guardar caché
    os.makedirs(os.path.dirname(CARDS_CACHE), exist_ok=True)
    with open(CARDS_CACHE, 'w', encoding='utf-8') as f:
        json.dump(cards, f, ensure_ascii=False, indent=2)
    print(f"  💾 Guardado en {os.path.basename(CARDS_CACHE)}")

    return cards


def load_cached_cards():
    """Carga las cartas desde el caché local si existe."""
    if not os.path.exists(CARDS_CACHE):
        return None
    with open(CARDS_CACHE, 'r', encoding='utf-8') as f:
        cards = json.load(f)
    return cards


def _normalize(text):
    """Normaliza un nombre para comparación."""
    import unicodedata
    text = unicodedata.normalize('NFD', text.lower())
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    return text


def _similarity(a, b):
    """Calcula similaridad entre dos strings normalizados."""
    return SequenceMatcher(None, _normalize(a), _normalize(b)).ratio()


def _name_match(excel_name, card_name):
    """
    Comprueba si el nombre del Excel coincide con la carta.
    Soporta nombres parciales (ej: "Nico Gonz" → "Nicolás González").
    """
    en = _normalize(excel_name)
    cn = _normalize(card_name)

    # Match exacto
    if en == cn:
        return 1.0

    # El nombre del excel es un prefijo o está contenido
    if cn.startswith(en) or en in cn:
        return 0.95

    # Todas las palabras del excel aparecen al inicio de alguna palabra de la carta
    excel_words = en.split()
    card_words = cn.split()
    if all(any(cw.startswith(ew) for cw in card_words) for ew in excel_words):
        return 0.9

    # Similaridad general
    return _similarity(excel_name, card_name)


def build_team_map(lineup, cards):
    """
    Para cada jugador del lineup, busca la carta más similar en la
    colección y devuelve un dict {nombre_excel: equipo}.
    """
    team_map = {}
    all_players = []
    for pos_data in lineup.values():
        all_players.extend(pos_data['in_season'])
        all_players.extend(pos_data['classic'])

    for excel_name, _ in all_players:
        if excel_name in team_map:
            continue

        best_score = 0
        best_team = '?'
        for card in cards:
            score = _name_match(excel_name, card['name'])
            if score > best_score:
                best_score = score
                best_team = card['team']
                if score >= 0.95:
                    break

        if best_score < 0.5:
            best_team = '?'

        team_map[excel_name] = best_team

    return team_map


# ============================================================
# CUOTAS DE APUESTAS (The Odds API)
# ============================================================

ODDS_TEAM_ALIASES = {
    'atletico madrid': ['atlético de madrid', 'atletico de madrid'],
    'athletic bilbao': ['athletic club'],
    'barcelona': ['fc barcelona'],
    'real betis': ['real betis balompié', 'real betis'],
    'celta vigo': ['rc celta de vigo', 'celta de vigo'],
    'espanyol': ['rcd espanyol'],
    'mallorca': ['rcd mallorca'],
    'getafe': ['getafe cf'],
    'girona': ['girona fc'],
    'villarreal': ['villarreal cf'],
    'valencia': ['valencia cf'],
    'sevilla': ['sevilla fc'],
    'osasuna': ['ca osasuna'],
    'alaves': ['deportivo alavés', 'deportivo alaves'],
    'las palmas': ['ud las palmas'],
    'leganes': ['cd leganés', 'cd leganes'],
    'valladolid': ['real valladolid'],
    'racing santander': ['racing de santander'],
}


def _read_odds_api_key():
    """Lee ODDS_API_KEY del config.txt."""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               '..', 'config', 'config.txt')
    if not os.path.exists(config_path):
        return None
    with open(config_path, 'r') as f:
        for line in f:
            if line.strip().startswith('ODDS_API_KEY='):
                key = line.strip().split('=', 1)[1].strip()
                return key if key else None
    return None


def fetch_odds():
    """
    Obtiene las cuotas de La Liga desde The Odds API.
    Cachea resultados 6 horas. Devuelve None si no hay API key.
    """
    api_key = _read_odds_api_key()
    if not api_key:
        return None

    # Comprobar caché (6 horas)
    if os.path.exists(ODDS_CACHE):
        try:
            with open(ODDS_CACHE, 'r') as f:
                cached = json.load(f)
            cached_time = datetime.fromisoformat(cached['timestamp'])
            age_hours = (datetime.now(timezone.utc) - cached_time).total_seconds() / 3600
            if age_hours < 6:
                print(f"📊 Usando cuotas cacheadas ({age_hours:.1f}h)")
                return cached['data']
        except (KeyError, ValueError):
            pass

    print("📊 Descargando cuotas de La Liga...")
    url = 'https://api.the-odds-api.com/v4/sports/soccer_spain_la_liga/odds/'
    params = {
        'apiKey': api_key,
        'regions': 'eu',
        'markets': 'h2h,totals',
        'oddsFormat': 'decimal',
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        os.makedirs(os.path.dirname(ODDS_CACHE), exist_ok=True)
        with open(ODDS_CACHE, 'w') as f:
            json.dump({
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'data': data,
            }, f, indent=2)
        remaining = resp.headers.get('x-requests-remaining', '?')
        print(f"   ✅ {len(data)} partidos  (requests restantes: {remaining})")
        return data
    except Exception as e:
        print(f"   ⚠️  Error obteniendo cuotas: {e}")
        return None


def _match_odds_team(odds_name, sorare_teams):
    """Busca el equipo Sorare que mejor coincide con el nombre de la API de cuotas."""
    on = _normalize(odds_name)

    # 1. Aliases manuales
    for alias_key, alias_vals in ODDS_TEAM_ALIASES.items():
        if on == _normalize(alias_key) or _normalize(alias_key) in on:
            for alias in alias_vals:
                na = _normalize(alias)
                # Primero buscar match exacto
                for st in sorare_teams:
                    if _normalize(st) == na:
                        return st
                # Luego buscar como substring
                for st in sorare_teams:
                    if na in _normalize(st):
                        return st
            return None  # Alias reconocido pero equipo no está en nuestra colección

    # 2. Fuzzy match directo
    best_score = 0
    best_team = None
    for st in sorare_teams:
        sim = _similarity(odds_name, st)
        if sim > best_score:
            best_score = sim
            best_team = st
    return best_team if best_score >= 0.45 else None


def build_odds_map(odds_data, sorare_teams):
    """
    Construye mapa de cuotas y lista de partidos.
    Filtra solo la jornada en curso (partidos agrupados en ~3 días).
    Returns: (odds_map, matches)
      odds_map: {sorare_team: {win_prob, opponent, home, avg_odds}}
      matches:  [{home, away, home_prob, away_prob, home_odds, away_odds}]
    """
    if not odds_data:
        return {}, []

    # Filtrar solo la jornada en curso: agrupar por fecha y tomar el primer bloque
    odds_with_time = []
    for m in odds_data:
        ct = m.get('commence_time', '')
        try:
            t = datetime.fromisoformat(ct.replace('Z', '+00:00'))
        except (ValueError, AttributeError):
            t = None
        odds_with_time.append((t, m))

    # Ordenar por fecha
    odds_with_time.sort(key=lambda x: x[0] or datetime.max.replace(tzinfo=timezone.utc))

    # Tomar solo los partidos dentro de 3 días desde el primero (una jornada típica)
    current_matchday = []
    first_time = None
    for t, m in odds_with_time:
        if t is None:
            continue
        if first_time is None:
            first_time = t
        if (t - first_time).total_seconds() <= 3 * 86400:  # 3 días
            current_matchday.append(m)
        else:
            break

    odds_map = {}
    matches = []

    for match in current_matchday:
        home = match['home_team']
        away = match['away_team']
        home_prices, away_prices, draw_prices = [], [], []
        over25_prices, under25_prices = [], []

        for bm in match.get('bookmakers', []):
            for mkt in bm.get('markets', []):
                if mkt['key'] == 'h2h':
                    for o in mkt['outcomes']:
                        if o['name'] == home:
                            home_prices.append(o['price'])
                        elif o['name'] == away:
                            away_prices.append(o['price'])
                        elif o['name'] == 'Draw':
                            draw_prices.append(o['price'])
                elif mkt['key'] == 'totals':
                    for o in mkt['outcomes']:
                        point = o.get('point', 2.5)
                        if point == 2.5:
                            if o['name'] == 'Over':
                                over25_prices.append(o['price'])
                            elif o['name'] == 'Under':
                                under25_prices.append(o['price'])

        if not home_prices or not away_prices:
            continue

        avg_h = sum(home_prices) / len(home_prices)
        avg_a = sum(away_prices) / len(away_prices)
        avg_d = sum(draw_prices) / len(draw_prices) if draw_prices else 3.5

        # Probabilidades implícitas normalizadas (sin overround)
        raw_h, raw_a, raw_d = 1 / avg_h, 1 / avg_a, 1 / avg_d
        total_raw = raw_h + raw_a + raw_d
        prob_h = raw_h / total_raw
        prob_a = raw_a / total_raw

        # Over 2.5 goles → probabilidad de partido con muchos goles
        if over25_prices and under25_prices:
            avg_over = sum(over25_prices) / len(over25_prices)
            avg_under = sum(under25_prices) / len(under25_prices)
            raw_o, raw_u = 1 / avg_over, 1 / avg_under
            over25_prob = raw_o / (raw_o + raw_u)
        else:
            over25_prob = 0.5  # neutral si no hay datos

        home_s = _match_odds_team(home, sorare_teams)
        away_s = _match_odds_team(away, sorare_teams)

        # Clean sheet estimado: pocos goles esperados + equipo favorito
        under25_prob = 1 - over25_prob
        cs_home = under25_prob * (0.4 + prob_h * 0.6)
        cs_away = under25_prob * (0.4 + prob_a * 0.6)

        if home_s:
            odds_map[home_s] = {
                'win_prob': prob_h, 'opponent': away_s or away,
                'home': True, 'avg_odds': avg_h,
                'over25': over25_prob,
                'cs_prob': cs_home,
            }
        if away_s:
            odds_map[away_s] = {
                'win_prob': prob_a, 'opponent': home_s or home,
                'home': False, 'avg_odds': avg_a,
                'over25': over25_prob,
                'cs_prob': cs_away,
            }

        matches.append({
            'home': home_s or home,
            'away': away_s or away,
            'home_prob': prob_h,
            'away_prob': prob_a,
            'home_odds': avg_h,
            'away_odds': avg_a,
            'over25': over25_prob,
            'commence': match.get('commence_time', ''),
        })

    return odds_map, matches


def print_matches(matches):
    """Imprime los partidos de la jornada con probabilidades."""
    if not matches:
        return
    print(f"\n{'─' * 78}")
    print(f"  ⚽ PARTIDOS DE LA JORNADA  ({len(matches)} partidos)")
    print(f"{'─' * 78}")
    for m in sorted(matches, key=lambda x: -max(x['home_prob'], x['away_prob'])):
        hp, ap = m['home_prob'] * 100, m['away_prob'] * 100
        o25 = m.get('over25', 0.5) * 100
        print(f"    {m['home']:<22} {hp:2.0f}% ({m['home_odds']:.2f})"
              f"  vs  ({m['away_odds']:.2f}) {ap:.0f}%  {m['away']}"
              f"   O2.5:{o25:.0f}%")
    print()


def _odds_factor(team, odds_map, pos=None):
    """
    Factor multiplicador por posición:
      POR/DEF: favorito + clean sheet prob alta = mejor
      MED: solo favorito
      DEL: favorito + over2.5 alto = mejor
    """
    if not odds_map or team not in odds_map:
        return 1.0
    info = odds_map[team]
    win_prob = info['win_prob']

    # Factor base: equipo favorito
    base = 1.0 + (win_prob - 0.33) * PESO_CUOTAS

    if pos in ('POR', 'DEF'):
        # Bonus por portería a 0 probable
        cs = info.get('cs_prob', 0.3)
        base += (cs - 0.3) * PESO_GOLES
    elif pos == 'DEL':
        # Bonus por partido con muchos goles
        over25 = info.get('over25', 0.5)
        base += (over25 - 0.5) * PESO_GOLES

    return base


# ============================================================
# OPTIMIZADOR DE ALINEACIONES
# ============================================================

POS_MAP = {'Porteros': 'POR', 'Defensas': 'DEF', 'Medios': 'MED', 'Delanteros': 'DEL'}
POS_LABELS = {'POR': 'POR', 'DEF': 'DEF', 'MED': 'MED', 'DEL': 'DEL'}
POS_ORDER = {'POR': 0, 'DEF': 1, 'MED': 2, 'DEL': 3}
# Distribuciones válidas de los 4 jugadores de campo (DEF, MED, DEL)
OUTFIELD_PATTERNS = [(2, 1, 1), (1, 2, 1), (1, 1, 2)]


def build_player_pool(lineup_data, team_map, odds_map=None):
    """Convierte los datos del Excel en lista plana con scores ajustados por cuotas."""
    players = []
    for pos_name, data in lineup_data.items():
        pos_code = POS_MAP[pos_name]
        for i, (name, score) in enumerate(data['in_season']):
            team = team_map.get(name, '?')
            factor = _odds_factor(team, odds_map, pos_code)
            players.append({
                'id': f"{pos_code}_IS_{i}",
                'name': name,
                'score': score,
                'adj_score': round(score * factor, 1),
                'pos': pos_code,
                'team': team,
                'classic': False,
            })
        for i, (name, score) in enumerate(data['classic']):
            team = team_map.get(name, '?')
            factor = _odds_factor(team, odds_map, pos_code)
            players.append({
                'id': f"{pos_code}_CL_{i}",
                'name': name,
                'score': score,
                'adj_score': round(score * factor, 1),
                'pos': pos_code,
                'team': team,
                'classic': True,
            })

    # Dedup: mismo nombre, puntos, posición y tipo = entrada duplicada en Excel
    seen = set()
    deduped = []
    for p in players:
        key = (p['name'], p['score'], p['pos'], p['classic'])
        if key not in seen:
            seen.add(key)
            deduped.append(p)
    return deduped


def generate_valid_lineups(players, max_score=260):
    """
    Genera alineaciones válidas de 5 jugadores.
    Filtro por score RAW ≤ max_score, ordenadas por score efectivo
    (adj_score + bonus POR+DEF mismo equipo).
    """
    gks = [p for p in players if p['pos'] == 'POR']
    defs = [p for p in players if p['pos'] == 'DEF']
    mids = [p for p in players if p['pos'] == 'MED']
    fwds = [p for p in players if p['pos'] == 'DEL']

    valid = []

    for gk in gks:
        for nd, nm, nf in OUTFIELD_PATTERNS:
            if len(defs) < nd or len(mids) < nm or len(fwds) < nf:
                continue
            for d_combo in combinations(defs, nd):
                for m_combo in combinations(mids, nm):
                    for f_combo in combinations(fwds, nf):
                        lineup = [gk] + list(d_combo) + list(m_combo) + list(f_combo)
                        raw_total = sum(p['score'] for p in lineup)
                        if raw_total > max_score:
                            continue
                        # Max 1 classic
                        if sum(1 for p in lineup if p['classic']) > 1:
                            continue
                        # Max 2 del mismo equipo
                        teams = Counter(p['team'] for p in lineup if p['team'] != '?')
                        if any(v > 2 for v in teams.values()):
                            continue
                        # Score efectivo = adj_scores + bonus POR+DEF
                        eff = sum(p['adj_score'] for p in lineup)
                        if gk['team'] != '?':
                            if any(p['pos'] == 'DEF' and p['team'] == gk['team']
                                   for p in lineup):
                                eff += BONUS_POR_DEF
                        valid.append((eff, lineup))

    valid.sort(key=lambda x: -x[0])
    return valid


def optimize_lineups(lineup_data, team_map, odds_map=None):
    """Encuentra las 4 mejores alineaciones disjuntas (maximiza score efectivo)."""
    players = build_player_pool(lineup_data, team_map, odds_map)
    has_odds = odds_map and any(p['adj_score'] != p['score'] for p in players)

    print(f"\n⚙️  Optimizando 4 alineaciones (máx {MAX_PUNTOS} pts cada una)...")
    if has_odds:
        print(f"   📊 Cuotas aplicadas (peso: {PESO_CUOTAS})")
    print(f"   Pool: {len(players)} jugadores "
          f"({sum(1 for p in players if p['pos']=='POR')} POR, "
          f"{sum(1 for p in players if p['pos']=='DEF')} DEF, "
          f"{sum(1 for p in players if p['pos']=='MED')} MED, "
          f"{sum(1 for p in players if p['pos']=='DEL')} DEL)")

    all_candidates = generate_valid_lineups(players, MAX_PUNTOS)
    print(f"   Combinaciones válidas: {len(all_candidates)}")

    if not all_candidates:
        print("❌ No se encontraron alineaciones válidas")
        return None

    best_result = None
    best_eff = 0

    # Probar las top N primeras elecciones, greedy para el resto
    top_1 = min(30, len(all_candidates))

    for i in range(top_1):
        eff_1, lineup_1 = all_candidates[i]
        used_1 = {p['id'] for p in lineup_1}
        rem_1 = [p for p in players if p['id'] not in used_1]

        cands_2 = generate_valid_lineups(rem_1, MAX_PUNTOS)
        if not cands_2:
            continue

        top_2 = min(10, len(cands_2))
        for j in range(top_2):
            eff_2, lineup_2 = cands_2[j]
            used_2 = used_1 | {p['id'] for p in lineup_2}
            rem_2 = [p for p in players if p['id'] not in used_2]

            cands_3 = generate_valid_lineups(rem_2, MAX_PUNTOS)
            if not cands_3:
                continue

            eff_3, lineup_3 = cands_3[0]
            used_3 = used_2 | {p['id'] for p in lineup_3}
            rem_3 = [p for p in players if p['id'] not in used_3]

            cands_4 = generate_valid_lineups(rem_3, MAX_PUNTOS)
            if not cands_4:
                continue

            eff_4, lineup_4 = cands_4[0]
            total_eff = eff_1 + eff_2 + eff_3 + eff_4

            if total_eff > best_eff:
                best_eff = total_eff
                best_result = [lineup_1, lineup_2, lineup_3, lineup_4]

    if best_result:
        raw_total = sum(p['score'] for lu in best_result for p in lu)
        if has_odds:
            print(f"   ✅ Mejor combinación: {raw_total} pts reales (eff: {best_eff:.0f})")
        else:
            print(f"   ✅ Mejor combinación encontrada: {raw_total} pts totales")
    else:
        print("   ❌ No se pudieron formar 4 alineaciones completas")

    return best_result


def print_lineups(lineups, odds_map=None):
    """Imprime las 4 alineaciones optimizadas."""
    total_all = 0
    for i, lineup in enumerate(lineups):
        lineup_sorted = sorted(lineup, key=lambda p: POS_ORDER.get(p['pos'], 99))
        total = sum(p['score'] for p in lineup)
        total_all += total
        classics = sum(1 for p in lineup if p['classic'])
        teams = Counter(p['team'] for p in lineup)
        dup_teams = [t for t, c in teams.items() if c > 1 and t != '?']

        # Detectar combo POR+DEF mismo equipo
        gk = next(p for p in lineup if p['pos'] == 'POR')
        gk_def = gk['team'] != '?' and any(
            p['pos'] == 'DEF' and p['team'] == gk['team'] for p in lineup)

        print(f"\n{'═' * 68}")
        header = f"  ALINEACIÓN {i+1}   │  {total} pts"
        if classics:
            header += "  │  1 classic"
        if dup_teams:
            header += f"  │  x2 {', '.join(dup_teams)}"
        if gk_def:
            header += "  │  🛡️POR+DEF"
        print(header)
        print(f"{'═' * 68}")

        for p in lineup_sorted:
            cl = ' ⭐' if p['classic'] else ''
            odds_info = ''
            if odds_map and p['team'] in odds_map:
                oi = odds_map[p['team']]
                loc = '🏠' if oi['home'] else '✈️ '
                odds_info = f"  {loc}{oi['win_prob']*100:3.0f}%"
                if p['pos'] in ('POR', 'DEF'):
                    cs = oi.get('cs_prob', 0)
                    odds_info += f" CS:{cs*100:.0f}%"
                elif p['pos'] == 'DEL':
                    o25 = oi.get('over25', 0)
                    odds_info += f" O2.5:{o25*100:.0f}%"
            print(f"    {POS_LABELS[p['pos']]}  {p['name']:<20} {p['score']:>3}"
                  f"   {p['team']}{cl}{odds_info}")

    print(f"\n{'━' * 68}")
    print(f"  TOTAL: {total_all} pts  (media: {total_all / len(lineups):.0f} pts/alineación)")
    print(f"{'━' * 68}")


def main():
    lineup = load_lineup()

    if ACTUALIZAR_CARTAS:
        print("\n🔄 Descargando cartas desde Sorare...")
        headers = build_headers()
        cards = fetch_my_rare_cards(headers)
    else:
        cards = load_cached_cards()
        if cards is None:
            print("\n📦 No hay caché de cartas, descargando por primera vez...")
            headers = build_headers()
            cards = fetch_my_rare_cards(headers)
        else:
            print(f"\n📦 Usando caché local ({len(cards)} cartas)")

    team_map = build_team_map(lineup, cards)

    # Obtener cuotas de apuestas
    odds_data = fetch_odds()
    all_known_teams = set(c['team'] for c in cards if c.get('team'))
    odds_map, matches = build_odds_map(odds_data, all_known_teams) if odds_data else ({}, [])
    if matches:
        print_matches(matches)
    elif PESO_CUOTAS > 0:
        print("\n⚠️  Sin cuotas. Añade ODDS_API_KEY en config/config.txt")
        print("   Consigue una gratis en: https://the-odds-api.com")

    if MOSTRAR_DETALLE:
        total_is = 0
        total_cl = 0
        for pos in POSITIONS:
            if pos not in lineup:
                continue
            data = lineup[pos]
            print_position(pos, data['in_season'], data['classic'], team_map)
            total_is += len(data['in_season'])
            total_cl += len(data['classic'])

        unmatched = [n for n, t in team_map.items() if t == '?']
        if unmatched:
            print(f"\n⚠️  Sin equipo encontrado: {', '.join(unmatched)}")

        print(f"\n{'═' * 55}")
        print(f"  RESUMEN")
        print(f"{'═' * 55}")
        print(f"  In-Season: {total_is} jugadores")
        print(f"  Classic:   {total_cl} jugadores")
        print(f"  Total:     {total_is + total_cl} jugadores")

    # Optimizar alineaciones
    result = optimize_lineups(lineup, team_map, odds_map)
    if result:
        print_lineups(result, odds_map)


if __name__ == '__main__':
    main()
