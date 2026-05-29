from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import subprocess
import sys
from typing import Any

from openpyxl import load_workbook

from .config_files import SorarePaths


@dataclass(frozen=True)
class CommandResult:
    command: str
    exit_code: int
    stdout: str
    stderr: str


@dataclass(frozen=True)
class SaleExecutionResult:
    ok: int
    fail: int
    skip: int
    items: list[dict[str, Any]]


def _run_command(cmd: list[str], cwd: Path, timeout: int = 1800) -> CommandResult:
    completed = subprocess.run(
        cmd,
        cwd=str(cwd),
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
    )
    return CommandResult(
        command=" ".join(cmd),
        exit_code=completed.returncode,
        stdout=completed.stdout.strip(),
        stderr=completed.stderr.strip(),
    )


def _normalize_rarity(rarity: str) -> str:
    key = (rarity or "").strip().lower()
    if key in {"super_rare", "azules"}:
        return "super_rare"
    if key in {"limited", "amarillas"}:
        return "limited"
    if key == "rare":
        return "rare"
    return "super_rare"


def excel_path_for_rarity(paths: SorarePaths, rarity: str) -> Path:
    normalized = _normalize_rarity(rarity)
    if normalized == "limited":
        suffix = "amarillas"
    elif normalized == "rare":
        suffix = "rojas"
    else:
        suffix = "azules"
    return paths.repo_root / "output" / f"cartas_para_vender_{suffix}.xlsx"


def run_export_cards(
    paths: SorarePaths,
    rarity: str,
    max_cards: int | None = None,
) -> CommandResult:
    normalized = _normalize_rarity(rarity)
    cmd = [sys.executable, str(paths.src_dir / "cartas_para_vender.py")]
    if normalized == "limited":
        cmd.append("--amarillas")
    elif normalized == "rare":
        cmd.append("--rojas")
    else:
        cmd.append("--azules")
    if max_cards is not None:
        cmd.extend(["--max-cartas", str(int(max_cards))])
    cmd.append("--no-open")
    return _run_command(cmd, paths.repo_root, timeout=2400)

def _header_map(ws) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            mapped[str(value).strip()] = col
    return mapped


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_sales_rows(excel_path: Path) -> list[dict[str, Any]]:
    if not excel_path.exists():
        return []

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    columns = _header_map(ws)

    if "Jugador" not in columns or "assetId" not in columns:
        return []

    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        player_name = ws.cell(row=row, column=columns["Jugador"]).value
        asset_id = ws.cell(row=row, column=columns["assetId"]).value
        if not player_name or not asset_id:
            continue

        row_data = {
            "row": row,
            "jugador": str(player_name),
            "asset_id": str(asset_id),
            "precio_venta": _to_float(ws.cell(row=row, column=columns.get("Precio venta (€)", 2)).value),
            "equipo": _read_cell(ws, row, columns, "Equipo"),
            "posicion": _read_cell(ws, row, columns, "Posición"),
            "nivel": _read_cell(ws, row, columns, "Nivel"),
            "temporada": _read_cell(ws, row, columns, "Temporada"),
            "liga": _read_cell(ws, row, columns, "Liga"),
            "coleccion": _read_cell(ws, row, columns, "Colección"),
            "rayos_coleccion": _to_float(_read_cell(ws, row, columns, "Rayos colección")),
            "rayos_carta": _to_float(_read_cell(ws, row, columns, "Rayos carta")) or 0,
            "rayos_tras_venta": _to_float(_read_cell(ws, row, columns, "Rayos tras venta")),
            "precio_medio": _to_float(_read_cell(ws, row, columns, "Precio Medio Ventas (€)")),
            "precio_min_classic": _to_float(_read_cell(ws, row, columns, "Precio Mín Classic (€)")),
            "precio_min_inseason": _to_float(_read_cell(ws, row, columns, "Precio Mín In Season (€)")),
            "in_season": str(_read_cell(ws, row, columns, "In Season") or "").strip().lower() in {"si", "sí", "yes", "true"},
        }
        rows.append(row_data)

    return rows


def _read_cell(ws, row: int, columns: dict[str, int], name: str):
    col = columns.get(name)
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def save_prices(excel_path: Path, request_data: dict[str, Any]) -> int:
    if not excel_path.exists():
        return 0

    wb = load_workbook(excel_path)
    ws = wb.active
    columns = _header_map(ws)
    price_col = columns.get("Precio venta (€)")
    if not price_col:
        return 0

    updates = 0
    for key, raw in request_data.items():
        if not key.startswith("price_row_"):
            continue
        try:
            row = int(key.replace("price_row_", ""))
        except ValueError:
            continue

        text = str(raw).strip()
        cell = ws.cell(row=row, column=price_col)
        if not text:
            cell.value = ""
            updates += 1
            continue

        value = _to_float(text)
        if value is None:
            continue
        cell.value = round(value, 2)
        updates += 1

    wb.save(excel_path)
    return updates


def reset_prices(excel_path: Path) -> int:
    if not excel_path.exists():
        return 0

    wb = load_workbook(excel_path)
    ws = wb.active
    columns = _header_map(ws)
    price_col = columns.get("Precio venta (€)")
    if not price_col:
        return 0

    cleared = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=price_col)
        if cell.value in (None, ""):
            continue
        cell.value = ""
        cleared += 1

    wb.save(excel_path)
    return cleared

def rows_ready_to_sell(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("precio_venta") is not None]


def execute_sales(paths: SorarePaths, rows: list[dict[str, Any]], selected_rows: list[int], days: int) -> SaleExecutionResult:
    js_script = paths.repo_root / "javascript" / "vender_carta.js"
    selected = {int(v) for v in selected_rows}
    _ = days
    sale_days = 2

    ok = 0
    fail = 0
    skip = 0
    items: list[dict[str, Any]] = []

    for row in rows:
        if row["row"] not in selected:
            continue

        price = row.get("precio_venta")
        if price is None:
            skip += 1
            items.append({"jugador": row["jugador"], "status": "skip", "message": "Sin precio de venta"})
            continue

        price_cents = int(round(float(price) * 100))
        cmd = ["node", str(js_script), row["asset_id"], str(price_cents), str(sale_days), "0"]
        result = _run_command(cmd, paths.repo_root, timeout=60)
        if result.exit_code == 0:
            ok += 1
            items.append({"jugador": row["jugador"], "status": "ok", "message": result.stdout or "Venta creada"})
        else:
            fail += 1
            items.append(
                {
                    "jugador": row["jugador"],
                    "status": "fail",
                    "message": result.stderr or result.stdout or "Error ejecutando venta",
                }
            )

    return SaleExecutionResult(ok=ok, fail=fail, skip=skip, items=items)
