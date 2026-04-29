#!/usr/bin/env python3
"""TelegramPriceAlert

Script de referencia para enviar una alerta a Telegram cuando el precio mínimo
(Buy Now / oferta activa más barata) de un jugador esté por debajo de un umbral.

Requisitos (en `config.txt`):
- JWT_TOKEN=...               (ya lo usas en el resto del repo)
- JWT_AUD=myapp               (opcional)

Telegram:
- TELEGRAM_BOT_TOKEN=123:ABC...
- TELEGRAM_CHAT_ID=123456789        (o @canal)

Uso:
- Ejecutar normal:  `.venv/bin/python src/TelegramPriceAlert.py`
- Solo previsualizar (sin enviar): `.venv/bin/python src/TelegramPriceAlert.py --dry-run`
- Demo (no llama a Sorare, usa 10.97€): `.venv/bin/python src/TelegramPriceAlert.py --demo --dry-run`
"""

import argparse
from dataclasses import dataclass
import os
import re
import unicodedata
import json
import time
from datetime import datetime
from typing import Optional

import requests

from openpyxl import load_workbook

from sorare_utils import (
    build_headers,
    fetch_exchange_rates,
    get_live_single_sale_offers,
    get_matching_offers,
    read_config,
    search_players_by_name,
    to_eur_cents,
)


SORARE_CARD_URL = "https://sorare.com/football/cards/{card_slug}"
SORARE_PLAYER_URL = "https://sorare.com/football/players/{player_slug}"


# ---------------------------------------------------------------------------
# Configuración de la alerta (ajusta aquí lo que quieras vigilar)
# ---------------------------------------------------------------------------

# Por defecto, el comportamiento se controla desde `telegram_alert_settings.txt`.
DEFAULT_SETTINGS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "telegram_alert_settings.txt")

# Defaults (si faltan en el settings file)
DEFAULT_NOTIFY_MODE = "all"  # all|edge|drop
DEFAULT_NOTIFY_DROP_EUR = 1.0
DEFAULT_SEND_ALL_OFFERS = True
DEFAULT_SEND_RUN_START_MESSAGE = True
DEFAULT_INCLUDE_PLAYER_PREVIEW = True
DEFAULT_RARITY = "rare"
DEFAULT_SEASON_YEAR: Optional[int] = None

DEFAULT_DESIRED_PLAYERS_FILE = "desired_players.txt"
DEFAULT_REFERENCE_EXCEL_FILE = os.path.join("output", "rare_cards.xlsx")
DEFAULT_STATE_FILE = os.path.join("output", "telegram_alert_state.json")


def _parse_bool(value: Optional[str], default: bool) -> bool:
    if value is None:
        return default
    v = value.strip().lower()
    if v in ("1", "true", "yes", "y", "on"):
        return True
    if v in ("0", "false", "no", "n", "off"):
        return False
    return default


def _parse_int_optional(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    v = value.strip()
    if v == "" or v.lower() == "none":
        return None
    return int(v)


def _parse_float(value: Optional[str], default: float) -> float:
    if value is None:
        return default
    v = value.strip().replace(",", ".")
    if v == "":
        return default
    return float(v)


@dataclass(frozen=True)
class CheapestOffer:
    price_eur: float
    card_slug: Optional[str]
    player_slug: str
    player_name: Optional[str]
    rarity: Optional[str]
    season_year: Optional[int]
    serial_number: Optional[int]
    grade: Optional[float]


def _normalize_text(value: str) -> str:
    value = value.strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"\s+", " ", value)
    return value


def _read_desired_players(path: str) -> list[tuple[str, float]]:
    desired: list[tuple[str, float]] = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue

            # Formato esperado: "nombre ... <numero>"
            parts = line.split()
            if len(parts) < 2:
                continue
            try:
                threshold = float(parts[-1].replace(",", "."))
            except ValueError as exc:
                raise SystemExit(
                    f"Línea inválida en {path}: '{line}'. Usa: <nombre jugador> <umbral_eur>"
                ) from exc
            name = " ".join(parts[:-1]).strip()
            if not name:
                continue
            desired.append((name, threshold))

    return desired


def _build_excel_index(path: str) -> list[tuple[str, str]]:
    wb = load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]

    if "assetId" not in header or "name" not in header:
        raise SystemExit(f"El Excel {path} no tiene columnas esperadas (name, assetId). Cabeceras: {header}")

    name_col = header.index("name") + 1
    asset_col = header.index("assetId") + 1

    rows: list[tuple[str, str]] = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        asset_id = ws.cell(row=row, column=asset_col).value
        if not name or not asset_id:
            continue
        rows.append((str(name), str(asset_id)))

    return rows


def _find_asset_id_for_player_name(excel_index: list[tuple[str, str]], player_name_query: str) -> Optional[str]:
    query_norm = _normalize_text(player_name_query)
    query_tokens = [t for t in query_norm.split(" ") if t]

    best_asset_id: Optional[str] = None
    best_score = -1

    for card_name, asset_id in excel_index:
        card_norm = _normalize_text(card_name)
        score = sum(1 for t in query_tokens if t in card_norm)
        if score > best_score and score > 0:
            best_score = score
            best_asset_id = asset_id
            if score == len(query_tokens):
                # match perfecto por tokens
                return best_asset_id

    return best_asset_id


def _pick_best_player_search_result(query_text: str, results: list[dict]) -> Optional[dict]:
    if not results:
        return None
    query_norm = _normalize_text(query_text)
    query_tokens = [t for t in query_norm.split(" ") if t]

    best = None
    best_score = -1
    for r in results:
        name = r.get('displayName') or ''
        slug = r.get('slug') or ''
        if not slug:
            continue
        name_norm = _normalize_text(name)
        score = sum(1 for t in query_tokens if t in name_norm)
        if score > best_score:
            best_score = score
            best = r
            if score == len(query_tokens):
                return best
    return best


def _cheapest_offer_for_player_slug(
    player_slug: str,
    *,
    rarity: str,
    season_year: Optional[int],
    rates,
    headers,
) -> Optional[CheapestOffer]:
    offers = get_live_single_sale_offers(player_slug, headers=headers)
    if not offers:
        return None

    cheapest = None
    cheapest_eur_cents = None

    for offer in offers:
        cards = offer.get('senderSide', {}).get('anyCards') or []
        if not cards:
            continue
        for c in cards:
            if rarity and str(c.get('rarityTyped', '')).lower() != str(rarity).lower():
                continue
            if season_year is not None and c.get('seasonYear') is not None and int(c['seasonYear']) != int(season_year):
                continue

            amounts = offer.get('receiverSide', {}).get('amounts')
            eur_cents = to_eur_cents(amounts, rates)
            if eur_cents is None:
                continue
            if cheapest_eur_cents is None or eur_cents < cheapest_eur_cents:
                cheapest_eur_cents = eur_cents
                player = (c.get('anyPlayer') or {})
                cheapest = CheapestOffer(
                    price_eur=float(eur_cents) / 100.0,
                    card_slug=c.get('slug'),
                    player_slug=player.get('slug') or player_slug,
                    player_name=player.get('displayName'),
                    rarity=c.get('rarityTyped'),
                    season_year=int(c.get('seasonYear')) if c.get('seasonYear') is not None else None,
                    serial_number=int(c.get('serialNumber')) if c.get('serialNumber') is not None else None,
                    grade=float(c.get('grade')) if c.get('grade') is not None else None,
                )

    return cheapest


def _offers_below_threshold_for_player_slug(
    player_slug: str,
    *,
    rarity: str,
    season_year: Optional[int],
    threshold_eur: float,
    rates,
    headers,
) -> list[CheapestOffer]:
    """Devuelve TODAS las cartas (ofertas activas) con precio < threshold."""
    offers = get_live_single_sale_offers(player_slug, headers=headers)
    if not offers:
        return []

    results: list[CheapestOffer] = []
    seen_card_slugs: set[str] = set()

    for offer in offers:
        cards = offer.get('senderSide', {}).get('anyCards') or []
        if not cards:
            continue

        amounts = offer.get('receiverSide', {}).get('amounts')
        eur_cents = to_eur_cents(amounts, rates)
        if eur_cents is None:
            continue
        price_eur = float(eur_cents) / 100.0
        if price_eur >= float(threshold_eur):
            continue

        for c in cards:
            if rarity and str(c.get('rarityTyped', '')).lower() != str(rarity).lower():
                continue
            if season_year is not None and c.get('seasonYear') is not None and int(c['seasonYear']) != int(season_year):
                continue

            card_slug = (c.get('slug') or '')
            if card_slug and card_slug in seen_card_slugs:
                continue
            if card_slug:
                seen_card_slugs.add(card_slug)

            player = (c.get('anyPlayer') or {})
            results.append(
                CheapestOffer(
                    price_eur=price_eur,
                    card_slug=c.get('slug'),
                    player_slug=player.get('slug') or player_slug,
                    player_name=player.get('displayName'),
                    rarity=c.get('rarityTyped'),
                    season_year=int(c.get('seasonYear')) if c.get('seasonYear') is not None else None,
                    serial_number=int(c.get('serialNumber')) if c.get('serialNumber') is not None else None,
                    grade=float(c.get('grade')) if c.get('grade') is not None else None,
                )
            )

    results.sort(key=lambda x: x.price_eur)
    return results


def _format_season_label(season_year: Optional[int]) -> Optional[str]:
    if season_year is None:
        return None
    return f"{season_year}-{str(season_year + 1)[-2:]}"


def _format_price_spanish(price_eur: float) -> str:
    # 10.95 -> "10,95"
    return f"{price_eur:.2f}".replace(".", ",")


def _format_level_from_grade(grade: Optional[float]) -> Optional[str]:
    if grade is None:
        return None
    # En Sorare suele ser un número entero en muchos casos (3.0, 7.0, ...)
    if abs(grade - round(grade)) < 1e-9:
        return str(int(round(grade)))
    return str(grade)


def _build_buy_link(offer: CheapestOffer) -> str:
    if offer.card_slug:
        return SORARE_CARD_URL.format(card_slug=offer.card_slug)
    return SORARE_PLAYER_URL.format(player_slug=offer.player_slug)


def _send_telegram_message(bot_token: str, chat_id: str, text: str) -> None:
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "parse_mode": "HTML",
        "disable_web_page_preview": False,
    }
    resp = requests.post(url, json=payload, timeout=30)
    resp.raise_for_status()


def _escape_html(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _build_run_start_message(desired_count: int) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return (
        "Sorare price alert\n"
        f"Ejecución: {now}\n"
        f"Jugadores: {desired_count}"
    )


def _build_alert_message(offer: CheapestOffer, *, include_preview_link: bool) -> str:
    """Construye el mensaje.

    - En Telegram: metemos el link del jugador *oculto* para que la preview enseñe su foto,
      pero visualmente solo mostramos el link de compra.
    - En consola (dry-run): no metemos el link oculto.
    """
    name = offer.player_name or offer.player_slug
    season_label = _format_season_label(offer.season_year) or ""
    level = _format_level_from_grade(offer.grade)
    price = _format_price_spanish(offer.price_eur)

    player_link = SORARE_PLAYER_URL.format(player_slug=offer.player_slug)
    buy_link = _build_buy_link(offer)

    # Texto visible
    visible_lines = [_escape_html(name)]
    if season_label:
        visible_lines.append(_escape_html(season_label))
    if level is not None:
        visible_lines.append(_escape_html(f"Nivel: {level}"))
    visible_lines.append(_escape_html(price))
    visible_lines.append(_escape_html(buy_link))
    visible_text = "\n".join(visible_lines)

    if not include_preview_link:
        return visible_text

    # Link oculto al principio para forzar preview del jugador
    hidden_preview = f"<a href=\"{_escape_html(player_link)}\">&#8205;</a>"
    return hidden_preview + visible_text


def _load_state(state_path: str) -> dict:
    if not os.path.isfile(state_path):
        return {}
    try:
        with open(state_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


def _save_state(state_path: str, state: dict) -> None:
    os.makedirs(os.path.dirname(state_path), exist_ok=True)
    with open(state_path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def main() -> int:
    parser = argparse.ArgumentParser(description="Alerta de bajada de precio a Telegram (Sorare)")
    parser.add_argument("--dry-run", action="store_true", help="No envía Telegram; solo imprime el mensaje")
    parser.add_argument("--demo", action="store_true", help="No llama a Sorare; usa un ejemplo (Militão 10.97€)")
    parser.add_argument(
        "--settings-file",
        default=DEFAULT_SETTINGS_PATH,
        help="Ruta al fichero de settings (por defecto: telegram_alert_settings.txt)",
    )
    parser.add_argument(
        "--desired-file",
        default=None,
        help="Ruta al fichero de jugadores deseados (por defecto: desired_players.txt)",
    )
    args = parser.parse_args()

    # config.txt (secrets: JWT + Telegram)
    secret_config = read_config()

    # settings file (comportamiento)
    settings = {}
    if args.settings_file and os.path.isfile(args.settings_file):
        settings = read_config(args.settings_file)

    config = {**secret_config, **settings}

    bot_token = config.get("TELEGRAM_BOT_TOKEN")
    chat_id = config.get("TELEGRAM_CHAT_ID")

    notify_mode = (config.get("NOTIFY_MODE") or DEFAULT_NOTIFY_MODE).strip().lower()
    notify_drop_eur = _parse_float(config.get("NOTIFY_DROP_EUR"), DEFAULT_NOTIFY_DROP_EUR)
    send_all_offers = _parse_bool(config.get("SEND_ALL_OFFERS_BELOW_THRESHOLD"), DEFAULT_SEND_ALL_OFFERS)
    send_run_start = _parse_bool(config.get("SEND_RUN_START_MESSAGE"), DEFAULT_SEND_RUN_START_MESSAGE)
    include_player_preview = _parse_bool(config.get("INCLUDE_PLAYER_PREVIEW"), DEFAULT_INCLUDE_PLAYER_PREVIEW)

    rarity = (config.get("RARITY") or DEFAULT_RARITY).strip()
    season_year = _parse_int_optional(config.get("SEASON_YEAR"))
    if season_year is None:
        season_year = DEFAULT_SEASON_YEAR

    desired_file = args.desired_file or (config.get("DESIRED_PLAYERS_FILE") or DEFAULT_DESIRED_PLAYERS_FILE)
    reference_excel_file = config.get("REFERENCE_EXCEL_FILE") or DEFAULT_REFERENCE_EXCEL_FILE
    state_file = config.get("STATE_FILE") or DEFAULT_STATE_FILE

    repo_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    desired_path = desired_file if os.path.isabs(desired_file) else os.path.join(repo_root, desired_file)
    reference_excel_path = (
        reference_excel_file if os.path.isabs(reference_excel_file) else os.path.join(repo_root, reference_excel_file)
    )
    state_path = state_file if os.path.isabs(state_file) else os.path.join(repo_root, state_file)

    if args.demo:
        offer = CheapestOffer(
            price_eur=10.97,
            card_slug=None,
            player_slug="eder-militao",
            player_name="Éder Militão",
            rarity="rare",
            season_year=2025,
            serial_number=47,
            grade=None,
        )
        message = _build_alert_message(offer, include_preview_link=(include_player_preview and (not args.dry_run)))
        print(message)
        if not args.dry_run:
            if not bot_token or not chat_id:
                raise SystemExit("Faltan TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID en config.txt")
            _send_telegram_message(bot_token, chat_id, message)
            print("✅ Mensaje enviado a Telegram (demo)")
        return 0

    if not bot_token or not chat_id:
        raise SystemExit("Faltan TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID en config.txt")

    headers = build_headers(config)
    rates = fetch_exchange_rates()

    if not os.path.isfile(desired_path):
        raise SystemExit(f"No existe el fichero de deseados: {desired_path}")
    if not os.path.isfile(reference_excel_path):
        raise SystemExit(
            f"No existe el Excel de referencia: {reference_excel_path}. "
            "Ejecuta CardsToExcel.py para generarlo."
        )

    desired = _read_desired_players(desired_path)
    if not desired:
        print("No hay jugadores en desired_players.txt")
        return 0

    # Mensaje de arranque por ejecución (solo en modo real)
    if (not args.dry_run) and send_run_start:
        _send_telegram_message(bot_token, chat_id, _build_run_start_message(len(desired)))
        print("✅ Inicio de ejecución enviado a Telegram")

    excel_index = _build_excel_index(reference_excel_path)

    state = _load_state(state_path)

    sent_keys: set[str] = set()

    sent = 0
    for desired_name, threshold_eur in desired:
        asset_id = _find_asset_id_for_player_name(excel_index, desired_name)
        offer: Optional[CheapestOffer] = None
        offers_below: list[CheapestOffer] = []
        player_name = None

        if asset_id:
            card, matching = get_matching_offers(asset_id, headers=headers, rates=rates)

            # Filtro por temporada si se ha definido
            if season_year is not None:
                matching = [m for m in matching if m.get("season") is not None and int(m["season"]) == int(season_year)]

            if matching:
                cheapest_match = matching[0]
                eur_cents = cheapest_match.get("sort_price")
                if eur_cents is None or eur_cents == float("inf"):
                    offer = None
                else:
                    player_slug = card["anyPlayer"]["slug"]
                    player_name = card["anyPlayer"].get("displayName")
                    offer = CheapestOffer(
                        price_eur=float(eur_cents) / 100.0,
                        card_slug=cheapest_match.get("slug"),
                        player_slug=player_slug,
                        player_name=player_name,
                        rarity=str(rarity) if rarity else None,
                        season_year=int(cheapest_match.get("season")) if cheapest_match.get("season") is not None else None,
                        serial_number=int(cheapest_match.get("serial")) if cheapest_match.get("serial") is not None else None,
                        grade=float(cheapest_match.get("grade")) if cheapest_match.get("grade") is not None else None,
                    )

                    # Calcula TODAS las ofertas bajo umbral usando `matching` (ya ordenado por precio)
                    seen = set()
                    for m in matching:
                        sp = m.get("sort_price")
                        if sp is None or sp == float("inf"):
                            continue
                        price = float(sp) / 100.0
                        if price >= float(threshold_eur):
                            continue
                        slug = (m.get("slug") or "")
                        if slug and slug in seen:
                            continue
                        if slug:
                            seen.add(slug)
                        offers_below.append(
                            CheapestOffer(
                                price_eur=price,
                                card_slug=m.get("slug"),
                                player_slug=player_slug,
                                player_name=player_name,
                                rarity=str(rarity) if rarity else None,
                                season_year=int(m.get("season")) if m.get("season") is not None else None,
                                serial_number=int(m.get("serial")) if m.get("serial") is not None else None,
                                grade=float(m.get("grade")) if m.get("grade") is not None else None,
                            )
                        )

        if not offer:
            # Fallback: buscar por nombre en Sorare y usar playerSlug real
            search_results = search_players_by_name(desired_name, headers=headers)
            best = _pick_best_player_search_result(desired_name, search_results)
            if not best:
                print(f"⚠️  No encontré jugador en Sorare para: {desired_name}")
                continue
            player_slug = best['slug']
            player_name = best.get('displayName')
            offer = _cheapest_offer_for_player_slug(
                player_slug,
                rarity=rarity,
                season_year=season_year,
                rates=rates,
                headers=headers,
            )
            if offer is None:
                print(f"{player_name or desired_name}: sin ofertas")
                continue

            offers_below = _offers_below_threshold_for_player_slug(
                player_slug,
                rarity=rarity,
                season_year=season_year,
                threshold_eur=float(threshold_eur),
                rates=rates,
                headers=headers,
            )

        # Deduplicación: no enviar 2 veces lo mismo (por doble ejecución accidental o duplicados)
        dedup_key = f"{offer.player_slug}|{offer.card_slug or ''}|{offer.price_eur:.2f}|{float(threshold_eur):.2f}"
        if dedup_key in sent_keys:
            continue
        sent_keys.add(dedup_key)

        previous = state.get(offer.player_slug) or {}
        was_below = bool(previous.get("below", False))
        below_now = offer.price_eur < float(threshold_eur)
        prev_threshold = previous.get("threshold_eur")
        if prev_threshold is not None and abs(float(prev_threshold) - float(threshold_eur)) > 1e-9:
            was_below = False

        previous_min = previous.get("last_min_price_eur")
        min_drop_trigger = False
        if was_below and below_now and previous_min is not None:
            try:
                min_drop_trigger = offer.price_eur <= (float(previous_min) - float(notify_drop_eur))
            except (ValueError, TypeError):
                min_drop_trigger = False

        should_notify = False
        if notify_mode == "all":
            should_notify = below_now
        elif notify_mode == "edge":
            should_notify = (below_now and (not was_below))
        elif notify_mode == "drop":
            should_notify = (below_now and ((not was_below) or min_drop_trigger))
        else:
            # fallback razonable
            should_notify = below_now

        if should_notify and below_now:
            to_send = offers_below if offers_below else [offer]
            # Dedupe por card_slug dentro del mismo jugador
            seen_send: set[str] = set()
            previously_sent_slugs = set(previous.get("sent_card_slugs") or [])

            # Si no estamos en modo 'all', evitamos repetir lo mismo mientras ya está en alerta:
            # - enviamos solo cartas nuevas
            # - o, si fue por drop-trigger, enviamos solo la más barata actual
            if notify_mode != "all" and (not args.dry_run) and was_below and not min_drop_trigger:
                to_send = [item for item in to_send if (item.card_slug or "") and (item.card_slug not in previously_sent_slugs)]
                if not to_send:
                    print(f"{player_name or desired_name}: (saltando) ya estaba en alerta")
                    continue

            if notify_mode == "drop" and (not args.dry_run) and was_below and min_drop_trigger:
                to_send = [offer]

            if not send_all_offers:
                to_send = [to_send[0]]

            for item in to_send:
                slug = item.card_slug or ""
                if slug and slug in seen_send:
                    continue
                if slug:
                    seen_send.add(slug)

                message = _build_alert_message(item, include_preview_link=(include_player_preview and (not args.dry_run)))
                print(message)
                if not args.dry_run:
                    _send_telegram_message(bot_token, chat_id, message)
                    sent += 1
                    print("✅ Mensaje enviado a Telegram")

            if not args.dry_run:
                sent_card_slugs = sorted(previously_sent_slugs.union(seen_send))
                state[offer.player_slug] = {
                    "card_slug": offer.card_slug or "",
                    "price_eur": float(f"{offer.price_eur:.2f}"),
                    "threshold_eur": float(f"{float(threshold_eur):.2f}"),
                    "below": True,
                    "sent_card_slugs": sent_card_slugs,
                    "last_min_price_eur": float(f"{offer.price_eur:.2f}"),
                    "sent_at": int(time.time()),
                }
        else:
            print(f"{player_name or desired_name}: sin alerta ({offer.price_eur:.2f}€ >= {threshold_eur:.2f}€)")
            if not args.dry_run:
                # Si sale de la zona de alerta, reseteamos el estado para poder avisar en la próxima bajada.
                state[offer.player_slug] = {
                    "card_slug": offer.card_slug or "",
                    "price_eur": float(f"{offer.price_eur:.2f}"),
                    "threshold_eur": float(f"{float(threshold_eur):.2f}"),
                    "below": False,
                    "sent_card_slugs": [],
                    "last_min_price_eur": None,
                    "updated_at": int(time.time()),
                }

    if sent == 0:
        print("No se enviaron alertas.")

    if not args.dry_run:
        _save_state(state_path, state)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
