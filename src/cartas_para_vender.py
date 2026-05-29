#!/usr/bin/env python3
"""
Genera un Excel con todas las cartas RARE, LIMITED o SUPER RARE que NO están en lineup,
incluyendo precio medio de últimas ventas y precio mínimo actual en mercado.

Uso:
  python cartas_para_vender.py              # rare (rojas)
  python cartas_para_vender.py --amarillas  # limited (amarillas)
    python cartas_para_vender.py --azules     # super_rare (azules)
"""
import os
import sys
import time
import json
import argparse
from collections import defaultdict
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, numbers

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sorare_utils import (
    graphql_request, build_headers, fetch_exchange_rates,
    get_min_prices_eur, get_recent_prices, get_card_info,
)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
DELAY = 0.2  # segundos entre llamadas a la API
MAX_CARTAS = 700  # Cuántas cartas consultar (pon un número grande para todas)
RAREZA = 'rojas'  # 'rojas' para rare, 'amarillas' para limited, 'azules' para super_rare


def fetch_cards_and_lineups(headers, rarity='rare'):
    """Descarga todas las cartas de la rareza indicada y la lista de slugs en lineup."""
    all_cards = []
    lineup_slugs = set()
    cursor = None
    page = 0

    while True:
        page += 1
        after_clause = f', after: "{cursor}"' if cursor else ''
        query = f"""
        query GetCardsPage {{
          currentUser {{
            cards(rarities: [{rarity}], first: 100{after_clause}) {{
              nodes {{
                assetId
                name
                slug
                rarityTyped
                seasonYear
                serialNumber
                anyPlayer {{
                  slug
                  displayName
                  activeClub {{
                    domesticLeague {{ name }}
                  }}
                }}
                anyTeam {{
                  name
                }}
                grade
                anyPositions
                inSeasonEligible
                tradeableStatus
                cardCollectionCards {{
                  scoreBreakdown {{
                    total
                    owner
                    holding
                    firstOwner
                    specialEdition
                    firstSerialNumber
                    shirtMatchingSerialNumber
                  }}
                  cardCollection {{ name }}
                }}
              }}
              pageInfo {{
                hasNextPage
                endCursor
              }}
            }}
            blockchainCardsInLineups
          }}
        }}
        """
        resp = requests.post(
            "https://api.sorare.com/graphql",
            headers=headers,
            json={"query": query},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if "errors" in data:
            raise RuntimeError(f"GraphQL error: {json.dumps(data['errors'])}")

        user = data["data"]["currentUser"]
        cards_data = user["cards"]
        all_cards.extend(cards_data["nodes"])
        lineup_slugs = set(user.get("blockchainCardsInLineups", []))

        if not cards_data["pageInfo"]["hasNextPage"]:
            break
        cursor = cards_data["pageInfo"]["endCursor"]
        print(f"\r  Cargando cartas... página {page}", end='', flush=True)

    print(f"\r  ✅ {len(all_cards)} cartas cargadas" + " " * 20)
    return all_cards, lineup_slugs


def build_collection_data(all_cards):
    """Agrupa cartas por colección.
    Devuelve:
      col_rayos:      {col_name: total_rayos}  (mejor carta por jugador)
      col_player_all: {(col_name, player_slug): [rayos, ...] desc}
    """
    col_player_all = defaultdict(list)
    for card in all_cards:
        player_slug = (card.get('anyPlayer') or {}).get('slug', card.get('slug', '?'))
        for ccc in card.get('cardCollectionCards', []):
            col_name = (ccc.get('cardCollection') or {}).get('name', '?')
            sb = ccc.get('scoreBreakdown') or {}
            rayos = sb.get('total', 0)
            col_player_all[(col_name, player_slug)].append(rayos)
    # Sort each list descending
    for key in col_player_all:
        col_player_all[key].sort(reverse=True)
    # Sum best per player
    col_rayos = defaultdict(int)
    for (col_name, player_slug), rayos_list in col_player_all.items():
        col_rayos[col_name] += rayos_list[0]
    return dict(col_rayos), dict(col_player_all)


def get_avg_recent_price(player_slug, rarity, season, headers, rates, cache):
    """Devuelve el precio medio de las últimas ventas en EUR, o None.
    Usa cache por (player_slug, rarity, season) para evitar duplicados."""
    cache_key = (player_slug, rarity, season)
    if cache_key in cache:
        return cache[cache_key]

    RETRY_DELAYS = [5, 15, 30]
    result = (None, 0)
    for attempt in range(len(RETRY_DELAYS) + 1):
        try:
            prices = get_recent_prices(player_slug, rarity, season=season, headers=headers)
            if not prices:
                result = (None, 0)
            else:
                eur_values = []
                for p in prices:
                    eur_cents = p.get('amounts', {}).get('eurCents')
                    if eur_cents:
                        eur_values.append(int(eur_cents) / 100)
                if not eur_values:
                    result = (None, 0)
                else:
                    result = (round(sum(eur_values) / len(eur_values), 2), len(eur_values))
            break  # éxito
        except Exception as e:
            if '429' in str(e) and attempt < len(RETRY_DELAYS):
                wait = RETRY_DELAYS[attempt]
                print(f"\n    ⏳ Rate limit (429), reintentando en {wait}s...")
                time.sleep(wait)
            else:
                print(f"\n    ⚠️  Error precios recientes: {e}")
                break

    cache[cache_key] = result
    return result


def get_min_price_cached(player_slug, rarity, asset_id, headers, rates, cache):
    """Precios mínimos en mercado (classic e in-season), cacheado por (player_slug, rarity).
    Devuelve (min_classic, min_inseason)."""
    cache_key = (player_slug, rarity)
    if cache_key in cache:
        return cache[cache_key]

    RETRY_DELAYS = [5, 15, 30]
    result = (None, None)
    for attempt in range(len(RETRY_DELAYS) + 1):
        try:
            result = get_min_prices_eur(asset_id, headers=headers, rates=rates)
            break  # éxito
        except Exception as e:
            if '429' in str(e) and attempt < len(RETRY_DELAYS):
                wait = RETRY_DELAYS[attempt]
                print(f"\n    ⏳ Rate limit (429), reintentando en {wait}s...")
                time.sleep(wait)
            else:
                print(f"\n    ⚠️  Error precios mínimos: {e}")
                break
    cache[cache_key] = result
    return result


def write_excel(cards_data, output_path, show_vault=False):
    """Genera el Excel con los datos de las cartas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartas para vender"

    # Cabeceras
    # Sin --vault: A=Jugador, B=Precio venta, C=Oferta mínima 90%, D=Precio Mín Classic, E=Precio Mín In Season,
    #              F=Equipo, G=Nivel, H=Temporada, I=Posición, J=Liga, K=In Season,
    #              L=Colección, M=Rayos colección, N=Rayos carta, O=Rayos tras venta,
    #              P=Precio Medio Ventas, Q=assetId
    # Con --vault: igual pero P=Precio Medio Ventas, Q=Vault, R=assetId
    headers_row = ['Jugador', 'Precio venta (€)', 'Oferta mínima 90%', 'Precio Mín Classic (€)', 'Precio Mín In Season (€)',
                   'Equipo', 'Nivel', 'Temporada', 'Posición', 'Liga', 'In Season',
                   'Colección', 'Rayos colección', 'Rayos carta', 'Rayos tras venta',
                   'Precio Medio Ventas (€)']
    if show_vault:
        headers_row.append('Vault')
    headers_row.append('assetId')
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers_row, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Datos
    for i, card in enumerate(cards_data, 2):
        ws.cell(row=i, column=1, value=card['name'])
        # col 2: Precio venta (vacía, el usuario la rellena)
        ws.cell(row=i, column=2, value='')
        # col 3: Oferta mínima (checkbox lógico en web)
        ws.cell(row=i, column=3, value='No')

        classic_cell = ws.cell(row=i, column=4)
        if card['min_price_classic'] is not None:
            classic_cell.value = card['min_price_classic']
            classic_cell.number_format = '#,##0.00 €'
        else:
            classic_cell.value = "Sin ofertas"

        inseason_cell = ws.cell(row=i, column=5)
        if card['min_price_inseason'] is not None:
            inseason_cell.value = card['min_price_inseason']
            inseason_cell.number_format = '#,##0.00 €'
        else:
            inseason_cell.value = "Sin ofertas"

        ws.cell(row=i, column=6, value=card['team'])
        ws.cell(row=i, column=7, value=card['grade'])
        ws.cell(row=i, column=8, value=card['season'])
        ws.cell(row=i, column=9, value=card['position'])
        ws.cell(row=i, column=10, value=card['league'])
        ws.cell(row=i, column=11, value='Sí' if card.get('in_season') else 'No')
        ws.cell(row=i, column=12, value=card['collection_name'])
        ws.cell(row=i, column=13, value=card['collection_rayos'])
        ws.cell(row=i, column=14, value=card['card_rayos'])
        ws.cell(row=i, column=15, value=card['rayos_col_after'])

        avg_cell = ws.cell(row=i, column=16)
        if card['avg_price'] is not None:
            avg_cell.value = card['avg_price']
            avg_cell.number_format = '#,##0.00 €'
        else:
            avg_cell.value = "Sin datos"

        # Color especial para cartas in-season: fondo verde claro en la celda del jugador
        if card.get('in_season'):
            ws.cell(row=i, column=1).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Columna "Vault" (sólo si show_vault)
        col_offset = 17
        if show_vault:
            vault_cell = ws.cell(row=i, column=col_offset, value='Sí' if card['in_vault'] else '')
            if card['in_vault']:
                vault_cell.font = Font(bold=True, color="FF0000")
            col_offset += 1

        ws.cell(row=i, column=col_offset, value=card['asset_id'])

    # Ajustar anchos
    if show_vault:
        widths = [25, 16, 16, 20, 20, 22, 8, 12, 14, 20, 10, 30, 16, 12, 16, 22, 8, 20]
    else:
        widths = [25, 16, 16, 20, 20, 22, 8, 12, 14, 20, 10, 30, 16, 12, 16, 22, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Autofiltro
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\n💾 Guardado en {os.path.basename(output_path)}")


def main():
    parser = argparse.ArgumentParser(description='Genera Excel de cartas para vender')
    parser.add_argument('--amarillas', action='store_true',
                        help='Generar Excel de cartas limited (amarillas) en vez de rare (rojas)')
    parser.add_argument('--rojas', action='store_true',
                        help='Generar Excel de cartas rare (rojas)')
    parser.add_argument('--azules', action='store_true',
                        help='Generar Excel de cartas super_rare (azules)')
    parser.add_argument('--vault', action='store_true',
                        help='Incluir también las cartas que están en vault (por defecto se excluyen)')
    parser.add_argument('--max-cartas', type=int, default=MAX_CARTAS,
                        help='Número máximo de cartas a exportar, ordenadas de más nuevas a más viejas')
    parser.add_argument('--no-open', action='store_true',
                        help='No abrir el Excel automáticamente al terminar')
    args = parser.parse_args()

    # CLI flags sobreescriben la constante RAREZA
    if args.azules:
        tipo = 'azules'
    elif args.amarillas:
        tipo = 'amarillas'
    elif args.rojas:
        tipo = 'rojas'
    else:
        tipo = RAREZA

    if tipo == 'azules':
        rarity = 'super_rare'
        suffix = '_azules'
        label = 'super_rare (azules)'
    elif tipo == 'amarillas':
        rarity = 'limited'
        suffix = '_amarillas'
        label = 'limited (amarillas)'
    else:
        rarity = 'rare'
        suffix = '_rojas'
        label = 'rare (rojas)'

    output_path = os.path.join(OUTPUT_DIR, f'cartas_para_vender{suffix}.xlsx')

    print(f"🔄 Conectando con Sorare... [{label}]")
    headers = build_headers()
    rates = fetch_exchange_rates()

    print(f"\n📥 Descargando cartas {label} y alineaciones...")
    all_cards, lineup_slugs = fetch_cards_and_lineups(headers, rarity=rarity)

    # Calcular rayos totales por colección
    collection_rayos, col_player_all = build_collection_data(all_cards)

    # Filtrar cartas NO en lineup
    available = [c for c in all_cards if c.get('slug') not in lineup_slugs]
    in_lineup = len(all_cards) - len(available)

    # Filtrar cartas en vault (a menos que se pida incluirlas con --vault)
    if not args.vault:
        before_vault = len(available)
        available = [c for c in available if c.get('tradeableStatus', 'YES') == 'YES']
        in_vault = before_vault - len(available)
        print(f"   {in_lineup} en lineup, {in_vault} en vault (excluidas), {len(available)} disponibles para vender")
    else:
        print(f"   {in_lineup} en lineup, {len(available)} disponibles para vender (vault incluido)")

    if not available:
        print("✅ No hay cartas disponibles fuera de lineup")
        return

    # Mantener orden del API (más recientes primero, igual que la interfaz de Sorare)

    # Limitar cantidad
    max_cards = int(args.max_cartas)
    if max_cards < len(available):
        print(f"   Limitado a {max_cards} cartas (de {len(available)})")
        available = available[:max_cards]

    # Consultar precios (con cache para no repetir llamadas)
    cards_data = []
    total = len(available)
    price_cache = {}  # (player_slug, rarity, season) → (avg_price, num_sales)
    min_price_cache = {}  # (player_slug, rarity) → min_price
    print(f"\n📊 Consultando precios de {total} cartas...")

    for i, card in enumerate(available):
        player = card.get('anyPlayer') or {}
        player_name = player.get('displayName', card.get('name', '?'))
        player_slug = player.get('slug', '')
        team = (card.get('anyTeam') or {}).get('name', '?')
        active_club = player.get('activeClub') or {}
        league = (active_club.get('domesticLeague') or {}).get('name', '?')
        season = card.get('seasonYear')
        rarity = card.get('rarityTyped', 'rare')
        asset_id = card.get('assetId', '')
        positions = card.get('anyPositions', [])
        pos_str = ', '.join(positions) if positions else '?'

        # Rayos de la carta y colección desde scoreBreakdown
        ccc_list = card.get('cardCollectionCards', [])
        if ccc_list:
            sb = ccc_list[0].get('scoreBreakdown') or {}
            card_rayos = sb.get('total', 0)
            collection_name = (ccc_list[0].get('cardCollection') or {}).get('name', '?')
        else:
            card_rayos = 0
            season_str = f"{season}-{str(season+1)[-2:]}" if season else '?'
            rarity_label = rarity.capitalize()
            collection_name = f"{team} {rarity_label} {season_str}"
        col_total_rayos = collection_rayos.get(collection_name, 0)

        # Calcular rayos tras venta (considerando duplicados del mismo jugador)
        all_player_rayos = col_player_all.get((collection_name, player_slug), [card_rayos])
        remaining = list(all_player_rayos)
        if card_rayos in remaining:
            remaining.remove(card_rayos)
        best_after = max(remaining) if remaining else 0
        delta = all_player_rayos[0] - best_after
        rayos_col_after = col_total_rayos - delta

        pct = (i + 1) / total * 100
        print(f"\r  [{i+1}/{total}] {pct:.0f}% - {player_name:<25}", end='', flush=True)

        # Precio medio últimas ventas (cacheado por jugador+rareza+temporada)
        avg_price, num_sales = get_avg_recent_price(
            player_slug, rarity, season, headers, rates, price_cache)
        time.sleep(DELAY)

        # Precio mínimo en mercado (cacheado por jugador+rareza)
        min_price_classic, min_price_inseason = get_min_price_cached(
            player_slug, rarity, asset_id, headers, rates, min_price_cache)
        time.sleep(DELAY)

        tradeable = card.get('tradeableStatus', 'YES')
        in_vault = tradeable != 'YES'
        in_season = card.get('inSeasonEligible', False)

        grade = card.get('grade', 0)

        season_str = f"{season}-{str(season+1)[-2:]}" if season else '?'
        cards_data.append({
            'name': player_name,
            'team': team,
            'grade': grade,
            'season': season_str,
            'position': pos_str,
            'league': league,
            'collection_name': collection_name,
            'collection_rayos': col_total_rayos,
            'card_rayos': card_rayos,
            'rayos_col_after': rayos_col_after,
            'avg_price': avg_price,
            'min_price_classic': min_price_classic,
            'min_price_inseason': min_price_inseason,
            'asset_id': asset_id,
            'in_vault': in_vault,
            'in_season': in_season,
        })

    print(f"\r  ✅ {total} cartas consultadas" + " " * 40)

    # Estadísticas rápidas
    with_avg = sum(1 for c in cards_data if c['avg_price'] is not None)
    with_min = sum(1 for c in cards_data if c['min_price_classic'] is not None or c['min_price_inseason'] is not None)
    in_season_count = sum(1 for c in cards_data if c['in_season'])
    print(f"   {with_avg} con historial de ventas, {with_min} con ofertas en mercado")
    print(f"   {in_season_count} cartas in-season (marcadas en verde en el Excel)")

    show_vault = args.vault

    write_excel(cards_data, output_path, show_vault=show_vault)

    # Abrir el Excel en macOS (opcional)
    if not args.no_open:
        import subprocess
        subprocess.Popen(['open', output_path])


if __name__ == '__main__':
    main()
