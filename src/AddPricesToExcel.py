#!/usr/bin/env python3
"""
Script para añadir una columna de precio mínimo (€) al Excel de rare_cards.
Lee el Excel generado por CardsToExcel.py, consulta el precio de mercado
para las N primeras cartas, y guarda el resultado.
"""

import os
import time
from openpyxl import load_workbook
from sorare_utils import build_headers, fetch_exchange_rates, get_min_price_eur

# ============================================================
# CONFIGURACIÓN
# ============================================================
# Número de cartas a las que consultar el precio (primeras N filas)
NUM_CARTAS_A_CONSULTAR = 10

# Ruta al Excel de entrada (generado por CardsToExcel.py)
EXCEL_INPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output', 'rare_cards.xlsx')

# Ruta al Excel de salida (se sobreescribe el mismo fichero; cambia si quieres otro)
EXCEL_OUTPUT = EXCEL_INPUT
# ============================================================


def main():
    # --- Validar que el Excel existe ---
    if not os.path.isfile(EXCEL_INPUT):
        print(f"Error: No se encontró el archivo Excel en {EXCEL_INPUT}")
        print("Ejecuta primero CardsToExcel.py para generarlo.")
        return

    # --- Preparar API ---
    headers = build_headers()

    print("Obteniendo tasas de cambio actuales...")
    rates = fetch_exchange_rates()
    usd_to_eur, gbp_to_eur, eth_to_eur = rates
    print(f"  1 USD = {usd_to_eur:.4f} EUR | 1 GBP = {gbp_to_eur:.4f} EUR | 1 ETH = {eth_to_eur:.2f} EUR\n")

    # --- Cargar Excel ---
    wb = load_workbook(EXCEL_INPUT)
    ws = wb.active

    # Detectar columnas por cabecera (fila 1)
    header_row = [cell.value for cell in ws[1]]
    asset_id_col = None
    for idx, h in enumerate(header_row):
        if h and 'assetid' in str(h).lower().replace('_', ''):
            asset_id_col = idx
            break

    if asset_id_col is None:
        print("Error: No se encontró la columna 'assetId' en el Excel.")
        return

    # Añadir (o localizar) columna "Precio Mínimo (€)"
    price_col_name = "Precio Mínimo (€)"
    if price_col_name in header_row:
        price_col_idx = header_row.index(price_col_name)
    else:
        price_col_idx = len(header_row)
        ws.cell(row=1, column=price_col_idx + 1, value=price_col_name)

    # --- Consultar precios ---
    total_rows = ws.max_row - 1  # sin cabecera
    rows_to_check = min(NUM_CARTAS_A_CONSULTAR, total_rows)

    print(f"Excel cargado: {total_rows} cartas en total")
    print(f"Consultando precio para las primeras {rows_to_check} cartas...\n")

    for i in range(2, rows_to_check + 2):  # fila 2 = primera fila de datos
        asset_id = ws.cell(row=i, column=asset_id_col + 1).value
        card_name = ws.cell(row=i, column=1).value or "?"

        if not asset_id:
            print(f"  [{i-1}/{rows_to_check}] {card_name} — sin assetId, saltando")
            ws.cell(row=i, column=price_col_idx + 1, value="N/A")
            continue

        print(f"  [{i-1}/{rows_to_check}] {card_name}...", end=" ", flush=True)
        min_price = get_min_price_eur(asset_id, headers=headers, rates=rates)

        if min_price is not None:
            ws.cell(row=i, column=price_col_idx + 1, value=round(min_price, 2))
            print(f"{min_price:.2f}€")
        else:
            ws.cell(row=i, column=price_col_idx + 1, value="Sin ofertas")
            print("Sin ofertas")

        # Pequeña pausa para no saturar la API
        time.sleep(0.5)

    # --- Guardar ---
    wb.save(EXCEL_OUTPUT)
    print(f"\nExcel guardado en: {EXCEL_OUTPUT}")
    print(f"Se actualizaron {rows_to_check} de {total_rows} filas con precios.")


if __name__ == '__main__':
    main()
