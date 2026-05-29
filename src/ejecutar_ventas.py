#!/usr/bin/env python3
"""
Lee output/cartas_para_vender_rojas.xlsx (o _amarillas/_azules),
coge las cartas que tienen un precio en la columna "Precio venta (€)" y ejecuta
la venta de cada una usando javascript/vender_carta.js.

Uso:
  python ejecutar_ventas.py              # rare (rojas)
  python ejecutar_ventas.py --amarillas  # limited (amarillas)
    python ejecutar_ventas.py --azules     # super_rare (azules)
"""
import os
import sys
import subprocess
import argparse
import openpyxl

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
JS_SCRIPT = os.path.join(BASE_DIR, 'javascript', 'vender_carta.js')
RAREZA = 'rojas'  # 'rojas' para rare, 'amarillas' para limited, 'azules' para super_rare

COL_MAP = {}  # se rellena dinámicamente desde la cabecera del Excel


def _build_col_map(ws):
    """Lee la fila de cabecera y mapea nombre → número de columna."""
    col_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            col_map[header.strip()] = col
    return col_map


def _col(name):
    """Devuelve el número de columna para un nombre de cabecera."""
    if name not in COL_MAP:
        return None
    return COL_MAP[name]


def main():
    parser = argparse.ArgumentParser(description='Ejecutar ventas desde Excel')
    parser.add_argument('--amarillas', action='store_true',
                        help='Usar Excel de cartas limited (amarillas) en vez de rare (rojas)')
    parser.add_argument('--rojas', action='store_true',
                        help='Usar Excel de cartas rare (rojas)')
    parser.add_argument('--azules', action='store_true',
                        help='Usar Excel de cartas super_rare (azules)')
    parser.add_argument('--dias', type=int, default=7,
                        help='Días que estará la carta a la venta (por defecto: 7)')
    args = parser.parse_args()

    # CLI flags sobreescriben la constante RAREZA
    if args.azules:
        tipo = 'azules'
    elif args.amarillas:
        tipo = 'amarillas'
    elif args.rojas:
        tipo = 'rojas'
    else:
        tipo = RAREZA

    suffix = f'_{tipo}'
    excel_path = os.path.join(BASE_DIR, 'output', f'cartas_para_vender{suffix}.xlsx')

    # Fallback al fichero antiguo sin sufijo si no existe el nuevo
    if not os.path.isfile(excel_path):
        legacy = os.path.join(BASE_DIR, 'output', 'cartas_para_vender.xlsx')
        if tipo == 'rojas' and os.path.isfile(legacy):
            excel_path = legacy
        else:
            print(f"❌ No se encuentra {excel_path}")
            sys.exit(1)

    if not os.path.isfile(JS_SCRIPT):
        print(f"❌ No se encuentra {JS_SCRIPT}")
        sys.exit(1)

    print(f"📂 Leyendo {os.path.basename(excel_path)}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    global COL_MAP
    COL_MAP = _build_col_map(ws)

    # Verificar columnas mínimas
    for required in ['Jugador', 'Precio venta (€)', 'assetId']:
        if required not in COL_MAP:
            print(f"❌ Columna '{required}' no encontrada en el Excel")
            print(f"   Columnas encontradas: {', '.join(COL_MAP.keys())}")
            sys.exit(1)

    ventas = []
    for row in range(2, ws.max_row + 1):
        precio = ws.cell(row=row, column=_col('Precio venta (€)')).value
        # Sólo procesar filas con un precio numérico en "Precio venta"
        if precio is None or str(precio).strip() == '':
            continue
        try:
            float(precio)
        except (ValueError, TypeError):
            continue
        asset_id = ws.cell(row=row, column=_col('assetId')).value
        nombre = ws.cell(row=row, column=_col('Jugador')).value or '?'

        if not asset_id:
            print(f"⚠️  Fila {row}: sin assetId, saltando")
            continue
        precio_eur = float(precio)

        def cell(header):
            c = _col(header)
            return ws.cell(row=row, column=c).value if c else None

        coleccion = cell('Colección') or '?'
        rayos_col = cell('Rayos colección') or 0
        rayos_carta = cell('Rayos carta') or 0
        rayos_after = cell('Rayos tras venta')
        equipo = cell('Equipo') or '?'
        temporada = cell('Temporada') or '?'
        liga = cell('Liga') or '?'
        nivel = cell('Nivel')
        posicion = cell('Posición') or '?'
        precio_medio = cell('Precio Medio Ventas (€)')
        precio_min_classic = cell('Precio Mín Classic (€)')
        precio_min_inseason = cell('Precio Mín In Season (€)')
        vault_val = str(cell('Vault') or '').strip()

        ventas.append({
            'row': row,
            'name': nombre,
            'asset_id': str(asset_id).strip(),
            'price_eur': precio_eur,
            'price_cents': int(precio_eur * 100),
            'coleccion': coleccion,
            'rayos_col': int(rayos_col) if rayos_col else 0,
            'rayos_carta': int(rayos_carta) if rayos_carta else 0,
            'rayos_after': int(rayos_after) if rayos_after is not None else None,
            'equipo': equipo,
            'temporada': temporada,
            'liga': liga,
            'nivel': nivel,
            'posicion': posicion,
            'precio_medio': precio_medio,
            'precio_min_classic': precio_min_classic,
            'precio_min_inseason': precio_min_inseason,
            'vault': vault_val.lower() in ('sí', 'si'),
        })

    if not ventas:
        print("ℹ️  No hay cartas con precio para vender")
        return

    print(f"🔍 {len(ventas)} carta(s) con precio para vender (duración: {args.dias} días)\n")

    ok = 0
    fail = 0
    skip = 0
    for idx, v in enumerate(ventas):
        print(f"\n{'='*60}")
        print(f"  📋 Carta {idx+1}/{len(ventas)}")
        print(f"{'='*60}")
        nivel_str = f"  [Nivel {v['nivel']}]" if v['nivel'] is not None else ''
        print(f"  🎴 {v['name']}{nivel_str}  ({v['equipo']}, {v['temporada']})")
        print(f"     Liga: {v['liga']}")
        print(f"     Posición: {v['posicion']}")
        print(f"     Colección: {v['coleccion']}")
        print(f"     Rayos carta: {v['rayos_carta']}")
        if v['rayos_after'] is not None:
            print(f"     Colección actual: {v['rayos_col']} rayos  →  tras vender: {v['rayos_after']} rayos")
        else:
            print(f"     Rayos colección: {v['rayos_col']}")
        if v['precio_medio'] is not None and isinstance(v['precio_medio'], (int, float)):
            print(f"     Precio medio ventas: {v['precio_medio']:.2f} €")
        if v['precio_min_classic'] is not None and isinstance(v['precio_min_classic'], (int, float)):
            print(f"     Precio mín classic:  {v['precio_min_classic']:.2f} €")
        if v['precio_min_inseason'] is not None and isinstance(v['precio_min_inseason'], (int, float)):
            print(f"     Precio mín in-season: {v['precio_min_inseason']:.2f} €")
        print(f"     💰 Precio de venta: {v['price_eur']:.2f} €")
        if v['vault']:
            print(f"     ⚠️  ESTA CARTA ESTÁ EN VAULT - no se puede vender")

        respuesta = input(f"\n  ¿Vender {v['name']} por {v['price_eur']:.2f} €? (s/n): ").strip().lower()
        if respuesta not in ('s', 'si', 'sí', 'y', 'yes'):
            print(f"  ⏭️  Saltada")
            skip += 1
            continue

        if v['vault']:
            print(f"  ⚠️  No se puede vender, está en vault. Saltando.")
            skip += 1
            continue

        print("  Vendiendo...")
        try:
            result = subprocess.run(
                ['node', JS_SCRIPT, v['asset_id'], str(v['price_cents']), str(args.dias)],
                capture_output=True, text=True, timeout=30,
            )
            if result.stdout:
                print(result.stdout)
            if result.returncode == 0:
                print(f"  ✅ Puesta a la venta")
                ok += 1
            else:
                print(f"  ❌ Error (código {result.returncode})")
                if result.stderr:
                    print(result.stderr)
                fail += 1
        except subprocess.TimeoutExpired:
            print("  ⏱️  Timeout")
            fail += 1
        except Exception as e:
            print(f"  ❌ {e}")
            fail += 1

    print(f"\n{'='*60}")
    print(f"📊 Resultado: {ok} puestas a la venta, {skip} saltadas, {fail} errores")


if __name__ == '__main__':
    main()
